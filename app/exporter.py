import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def export_experiment_excel(label, spectrum):
    """Export data spektrum percobaan ke file Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Spektrum"

    # Styling
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="center")

    # Title row
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = label
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    # Info
    ws["A2"] = "Mode: Capture gambar tunggal"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)

    # Header row
    headers = ["Panjang Gelombang (nm)", "Intensitas Cahaya (Lux)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, point in enumerate(spectrum):
        row = i + 5
        wl_cell = ws.cell(row=row, column=1, value=point["wl"])
        lux_cell = ws.cell(row=row, column=2, value=point["lux"])
        for cell in [wl_cell, lux_cell]:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_absorbance_excel(result):
    """Export hasil analisis absorbansi ke Excel."""
    wb = Workbook()

    # --- Sheet 1: Absorbansi ---
    ws1 = wb.active
    ws1.title = "Absorbansi"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="center")

    ws1.merge_cells("A1:D1")
    ws1["A1"].value = "Hasil Analisis Absorbansi"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    ws1["A2"].value = "A(\u03bb) = -log10(I(\u03bb) / I\u2080(\u03bb))"
    ws1["A2"].font = Font(name="Calibri", size=10, italic=True)

    headers = ["Panjang Gelombang (nm)", "I\u2080 Aquadest (Lux)", "I Sampel (Lux)", "Absorbansi (A)"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    abs_data = result["absorbance"]
    ref = result["avg_reference"]
    sample = result["avg_sample"]

    for i in range(len(abs_data)):
        row = i + 5
        cells = [
            ws1.cell(row=row, column=1, value=abs_data[i]["wl"]),
            ws1.cell(row=row, column=2, value=ref[i]["lux"] if i < len(ref) else 0),
            ws1.cell(row=row, column=3, value=sample[i]["lux"] if i < len(sample) else 0),
            ws1.cell(row=row, column=4, value=abs_data[i]["A"]),
        ]
        for c in cells:
            c.font = data_font
            c.alignment = data_align
            c.border = thin_border

    for col_letter in ["A", "B", "C", "D"]:
        ws1.column_dimensions[col_letter].width = 26

    # --- Sheet 2: Data Mentah Per Grup ---
    ws2 = wb.create_sheet("Data Mentah Per Grup")
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Rata-rata Intensitas Per Grup"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    raw_headers = ["Panjang Gelombang (nm)", "Kuvet Kosong (Lux)", "Aquadest I\u2080 (Lux)", "Larutan Sampel I (Lux)"]
    for col, h in enumerate(raw_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    blank = result["avg_blank"]
    for i in range(len(ref)):
        row = i + 4
        cells = [
            ws2.cell(row=row, column=1, value=ref[i]["wl"]),
            ws2.cell(row=row, column=2, value=blank[i]["lux"] if i < len(blank) else 0),
            ws2.cell(row=row, column=3, value=ref[i]["lux"]),
            ws2.cell(row=row, column=4, value=sample[i]["lux"] if i < len(sample) else 0),
        ]
        for c in cells:
            c.font = data_font
            c.alignment = data_align
            c.border = thin_border

    for col_letter in ["A", "B", "C", "D"]:
        ws2.column_dimensions[col_letter].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
